from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class Scenario:
    sheet_name: str
    scenario_name: str
    size: int
    workers: list[str]
    machines: list[str]
    cost_matrix: list[list[int]]


def _is_scenario_label(value: object) -> bool:
    if not isinstance(value, str):
        return False
    stripped = value.strip()
    if not stripped:
        return False
    return "pekerja" not in stripped.lower()


def load_scenarios(dataset_path: Path) -> list[Scenario]:
    workbook = load_workbook(dataset_path, data_only=True)
    scenarios: list[Scenario] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        row_index = 0
        while row_index < len(rows):
            current = rows[row_index][0] if rows[row_index] else None
            if not _is_scenario_label(current):
                row_index += 1
                continue

            scenario_name = str(current).strip()
            header_row = rows[row_index + 1]
            machines = [str(value) for value in header_row[1:] if value is not None]
            size = len(machines)

            workers: list[str] = []
            matrix: list[list[int]] = []
            for offset in range(size):
                data_row = rows[row_index + 2 + offset]
                workers.append(str(data_row[0]))
                matrix.append([int(value) for value in data_row[1 : size + 1]])

            scenarios.append(
                Scenario(
                    sheet_name=worksheet.title,
                    scenario_name=scenario_name,
                    size=size,
                    workers=workers,
                    machines=machines,
                    cost_matrix=matrix,
                )
            )
            row_index += size + 2

    return scenarios


def iter_scenarios(dataset_path: Path) -> Iterable[Scenario]:
    yield from load_scenarios(dataset_path)